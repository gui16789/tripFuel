from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests
import uvicorn
import openpyxl
from fastapi import FastAPI, File, Header, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from generate_vehicle_usage import Trip, populate_usage_sheet, write_usage_sheet
from oil_price_fetcher import merge_and_save_prices, scrape_jilin_prices


BASE_DIR = Path(__file__).resolve().parent
WEB_DIR = BASE_DIR / "web"
FUEL_PRICE_PATH = BASE_DIR / "fuel_prices_changchun_92.csv"
TRIPS_PATH = BASE_DIR / "vehicle_usage_editor_records.json"
FUEL_DETAILS_DRAFT_PATH = BASE_DIR / "fuel_details_editor_records.json"
DESTINATION_POOL_PATH = BASE_DIR / "destination_pool.json"
PRIVATE_WORKBOOK = BASE_DIR / "加油明细.xlsx"
TEMPLATE_WORKBOOK = BASE_DIR / "templates" / "加油明细模板.xlsx"
CONFIGURED_WORKBOOK = Path(os.environ["SOURCE_WORKBOOK"]) if os.environ.get("SOURCE_WORKBOOK") else None
DEFAULT_OUTPUT = BASE_DIR / "加油明细_在线编辑器生成.xlsx"

ORIGIN_NAME = "安盟财产保险有限公司长春中心支公司"
CITY = "长春"
FUEL_RATE = 0.09
AMAP_BASE = "https://restapi.amap.com/v3"


app = FastAPI(title="车辆使用明细表编辑器")


class PoiSearchRequest(BaseModel):
    keyword: str
    city: str = CITY


class RouteRequest(BaseModel):
    origin: str = ORIGIN_NAME
    stops: list[dict[str, Any]]


class FuelPriceRefreshRequest(BaseModel):
    year_month: str = "2026-04"


class ExportRequest(BaseModel):
    records: list[dict[str, Any]]
    output_name: str | None = None


class CombinedExportRequest(BaseModel):
    records: list[dict[str, Any]]
    fuel_details: list[dict[str, Any]]
    output_name: str | None = None


class SaveRecordsRequest(BaseModel):
    records: list[dict[str, Any]]


class FuelDetailsRequest(BaseModel):
    rows: list[dict[str, Any]]
    output_name: str | None = None


class DestinationPoolRequest(BaseModel):
    destinations: list[dict[str, Any]]


def active_workbook_path() -> Path:
    candidates = [
        CONFIGURED_WORKBOOK,
        PRIVATE_WORKBOOK,
        TEMPLATE_WORKBOOK,
    ]
    for path in candidates:
        if path and path.exists():
            return path
    return PRIVATE_WORKBOOK


def is_template_workbook(path: Path) -> bool:
    return path.exists() and TEMPLATE_WORKBOOK.exists() and path.resolve() == TEMPLATE_WORKBOOK.resolve()


def workbook_payload(rows: list[dict[str, Any]], source: str, workbook_path: Path) -> dict[str, Any]:
    return {
        "rows": rows,
        "source": source,
        "workbook": str(workbook_path),
        "is_template": is_template_workbook(workbook_path),
    }


def excel_download_response(stream: BytesIO, filename: str) -> StreamingResponse:
    stream.seek(0)
    safe_filename = Path(filename).name
    if not safe_filename.lower().endswith(".xlsx"):
        safe_filename = f"{safe_filename}.xlsx"
    quoted = quote(safe_filename)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted}"},
    )


def amap_key(x_amap_key: str | None = Header(default=None)) -> str:
    key = x_amap_key or os.environ.get("AMAP_KEY", "")
    if not key:
        raise HTTPException(status_code=400, detail="缺少高德 AMAP_KEY，请在页面输入或设置环境变量。")
    return key


def amap_get(path: str, key: str, params: dict[str, Any]) -> dict[str, Any]:
    request_params = {"key": key, "output": "JSON", **params}
    response = requests.get(
        f"{AMAP_BASE}/{path}",
        params=request_params,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=20,
    )
    response.raise_for_status()
    data = response.json()
    if str(data.get("status")) != "1":
        raise HTTPException(status_code=502, detail=f"高德接口失败：{data.get('info') or data}")
    return data


def driving_distance(origin_location: str, destination_location: str, key: str) -> dict[str, Any]:
    data = amap_get(
        "direction/driving",
        key,
        {
            "origin": origin_location,
            "destination": destination_location,
            "strategy": 0,
            "extensions": "base",
        },
    )
    paths = (data.get("route") or {}).get("paths") or []
    if not paths:
        raise HTTPException(status_code=502, detail="高德没有返回可用驾车路线。")
    distance_m = int(float(paths[0]["distance"]))
    duration_s = int(float(paths[0].get("duration") or 0))
    return {
        "distance_m": distance_m,
        "distance_km": round(distance_m / 1000),
        "distance_km_exact": round(distance_m / 1000, 2),
        "duration_min": round(duration_s / 60, 1),
        "duration_s": duration_s,
    }


def geocode(address: str, key: str, city: str = CITY) -> dict[str, Any]:
    data = amap_get("geocode/geo", key, {"address": address, "city": city})
    geocodes = data.get("geocodes") or []
    if not geocodes:
        raise HTTPException(status_code=404, detail=f"无法解析地址：{address}")
    first = geocodes[0]
    return {
        "name": address,
        "address": first.get("formatted_address") or address,
        "location": first.get("location"),
        "district": first.get("district") or "",
        "level": first.get("level") or "",
    }


def read_fuel_prices() -> list[dict[str, Any]]:
    if not FUEL_PRICE_PATH.exists():
        return []
    with FUEL_PRICE_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def price_for_date(date_text: str) -> dict[str, Any] | None:
    try:
        trip_date = datetime.strptime(date_text, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="日期格式需要是 YYYY-MM-DD") from exc

    candidates = []
    for row in read_fuel_prices():
        effective = datetime.strptime(row["effective_date"], "%Y-%m-%d")
        if effective <= trip_date and row.get("fuel_type") == "92":
            candidates.append((effective, row))

    if not candidates:
        return None
    row = sorted(candidates, key=lambda item: item[0])[-1][1]
    return {
        "effective_date": row["effective_date"],
        "price": float(row["price"]),
        "source_url": row.get("source_url", ""),
        "fuel_type": row.get("fuel_type", "92"),
    }


def read_fuel_details_from_workbook(workbook_source: Any | None = None) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(workbook_source or active_workbook_path(), data_only=True)
    ws = workbook["加油明细"]
    rows: list[dict[str, Any]] = []

    for row in range(3, ws.max_row + 1):
        date = ws.cell(row, 1).value
        vehicle = ws.cell(row, 2).value
        driver = ws.cell(row, 3).value
        amount = ws.cell(row, 4).value
        liters = ws.cell(row, 5).value
        if str(date).strip() == "合计":
            break
        if not any(value is not None for value in [date, vehicle, driver, amount, liters]):
            continue
        if not isinstance(date, datetime):
            continue
        rows.append(
            {
                "date": date.strftime("%Y-%m-%d"),
                "vehicle": vehicle or "",
                "driver": driver or "",
                "amount": float(amount or 0),
                "liters": float(liters or 0),
            }
        )
    return rows


def populate_fuel_details_sheet(workbook: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    ws = workbook["加油明细"]

    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 3:
            ws.unmerge_cells(str(merged))

    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    for row_index, item in enumerate(rows, start=3):
        date = datetime.strptime(item["date"], "%Y-%m-%d")
        ws.cell(row_index, 1, date)
        ws.cell(row_index, 1).number_format = "yyyy-mm-dd"
        ws.cell(row_index, 2, item.get("vehicle") or "")
        ws.cell(row_index, 3, item.get("driver") or "")
        ws.cell(row_index, 4, float(item.get("amount") or 0))
        ws.cell(row_index, 5, float(item.get("liters") or 0))

    total_row = 3 + len(rows)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 4, f"=SUM(D3:D{total_row - 1})" if rows else 0)
    ws.cell(total_row, 5, f"=SUM(E3:E{total_row - 1})" if rows else 0)


def write_fuel_details(output_target: Any, rows: list[dict[str, Any]], workbook_path: Path | None = None) -> None:
    workbook = openpyxl.load_workbook(workbook_path or active_workbook_path())
    populate_fuel_details_sheet(workbook, rows)

    if isinstance(output_target, (str, Path)):
        output_path = Path(output_target)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    else:
        workbook.save(output_target)


def trips_from_records(records: list[dict[str, Any]]) -> list[Trip]:
    return [
        Trip(
            date=datetime.strptime(item["date"], "%Y-%m-%d"),
            origin=item.get("origin") or ORIGIN_NAME,
            destination=item.get("poi_name") or item.get("destination") or "",
            fuel_price=float(item["fuel_price"]),
            distance_km=float(item["distance_km"]),
        )
        for item in records
    ]


def default_destination_pool() -> list[dict[str, Any]]:
    return []


def read_destination_pool() -> list[dict[str, Any]]:
    if DESTINATION_POOL_PATH.exists():
        data = json.loads(DESTINATION_POOL_PATH.read_text(encoding="utf-8"))
        return data.get("destinations", [])
    return default_destination_pool()


def save_destination_pool(destinations: list[dict[str, Any]]) -> None:
    DESTINATION_POOL_PATH.write_text(
        json.dumps({"destinations": destinations}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


@app.get("/")
def index() -> FileResponse:
    return FileResponse(WEB_DIR / "index.html")


@app.get("/api/config")
def config() -> dict[str, Any]:
    workbook_path = active_workbook_path()
    return {
        "origin": ORIGIN_NAME,
        "city": CITY,
        "fuel_rate": FUEL_RATE,
        "has_server_key": bool(os.environ.get("AMAP_KEY")),
        "source_workbook": str(workbook_path),
        "is_template_workbook": is_template_workbook(workbook_path),
    }


@app.get("/api/fuel-prices")
def fuel_prices() -> dict[str, Any]:
    return {"prices": read_fuel_prices()}


@app.get("/api/fuel-price")
def fuel_price(date: str) -> dict[str, Any]:
    price = price_for_date(date)
    return {"price": price}


@app.post("/api/fuel-prices/refresh")
def refresh_fuel_prices(payload: FuelPriceRefreshRequest) -> dict[str, Any]:
    prices = scrape_jilin_prices(payload.year_month)
    merged = merge_and_save_prices(prices, FUEL_PRICE_PATH)
    return {"prices": [item.__dict__ for item in merged], "refreshed": [item.__dict__ for item in prices]}


@app.post("/api/pois")
def pois(payload: PoiSearchRequest, x_amap_key: str | None = Header(default=None)) -> dict[str, Any]:
    key = amap_key(x_amap_key)
    keyword = payload.keyword.strip()
    if not keyword:
        return {"pois": []}

    data = amap_get(
        "place/text",
        key,
        {
            "keywords": keyword,
            "city": payload.city,
            "citylimit": "false",
            "offset": 12,
            "page": 1,
            "extensions": "base",
        },
    )
    pois_data = []
    for item in data.get("pois") or []:
        location = item.get("location")
        if not location:
            continue
        pois_data.append(
            {
                "name": item.get("name") or keyword,
                "address": item.get("address") if isinstance(item.get("address"), str) else "",
                "district": item.get("adname") or "",
                "type": item.get("type") or "",
                "location": location,
            }
        )

    if not pois_data:
        fallback = geocode(keyword, key, payload.city)
        pois_data.append({**fallback, "type": "地理编码"})
    return {"pois": pois_data}


@app.post("/api/route")
def route(payload: RouteRequest, x_amap_key: str | None = Header(default=None)) -> dict[str, Any]:
    key = amap_key(x_amap_key)
    if not payload.stops:
        raise HTTPException(status_code=400, detail="至少需要一个目的地。")

    origin_location = geocode(payload.origin, key)["location"]
    locations = [stop.get("location") for stop in payload.stops if stop.get("location")]
    if len(locations) != len(payload.stops):
        raise HTTPException(status_code=400, detail="目的地缺少经纬度，请先选择高德候选地点。")

    params: dict[str, Any] = {
        "origin": origin_location,
        "destination": locations[-1],
        "strategy": 0,
        "extensions": "base",
    }
    if len(locations) > 1:
        params["waypoints"] = ";".join(locations[:-1])

    data = amap_get("direction/driving", key, params)
    paths = (data.get("route") or {}).get("paths") or []
    if not paths:
        raise HTTPException(status_code=502, detail="高德没有返回可用驾车路线。")

    outbound_distance_m = int(float(paths[0]["distance"]))
    duration_s = int(float(paths[0].get("duration") or 0))
    steps = [
        {
            "distance_m": int(float(step.get("distance") or 0)),
            "instruction": step.get("instruction") or "",
            "road": step.get("road") or "",
        }
        for step in paths[0].get("steps") or []
    ]
    outbound_segments = [
        {
            **driving_distance(
                origin_location if index == 0 else payload.stops[index - 1].get("location"),
                stop.get("location"),
                key,
            ),
            "from": payload.origin if index == 0 else payload.stops[index - 1].get("name"),
            "to": stop.get("name"),
            "location": stop.get("location"),
        }
        for index, stop in enumerate(payload.stops)
    ]
    return_segment = {
        **driving_distance(locations[-1], origin_location, key),
        "from": payload.stops[-1].get("name"),
        "to": payload.origin,
        "location": origin_location,
        "return_trip": True,
    }
    distance_m = outbound_distance_m + int(return_segment["distance_m"])
    duration_s += int(return_segment.get("duration_s") or 0)
    return {
        "distance_m": distance_m,
        "distance_km": round(distance_m / 1000),
        "distance_km_exact": round(distance_m / 1000, 2),
        "duration_min": round(duration_s / 60, 1),
        "outbound_distance_m": outbound_distance_m,
        "return_distance_m": return_segment["distance_m"],
        "round_trip": True,
        "segments": [*outbound_segments, return_segment],
        "steps": steps,
    }


@app.get("/api/records")
def load_records() -> dict[str, Any]:
    if not TRIPS_PATH.exists():
        return {"records": []}
    return json.loads(TRIPS_PATH.read_text(encoding="utf-8"))


@app.post("/api/records")
def save_records(payload: SaveRecordsRequest) -> dict[str, Any]:
    data = {"records": payload.records}
    TRIPS_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"saved": True, "count": len(payload.records)}


@app.get("/api/fuel-details")
def fuel_details() -> dict[str, Any]:
    if FUEL_DETAILS_DRAFT_PATH.exists():
        data = json.loads(FUEL_DETAILS_DRAFT_PATH.read_text(encoding="utf-8"))
        return {**data, "source": "draft", "workbook": str(active_workbook_path())}
    workbook_path = active_workbook_path()
    return workbook_payload(read_fuel_details_from_workbook(workbook_path), "workbook", workbook_path)


@app.get("/api/fuel-details/source")
def fuel_details_source() -> dict[str, Any]:
    workbook_path = active_workbook_path()
    return workbook_payload(read_fuel_details_from_workbook(workbook_path), "workbook", workbook_path)


@app.post("/api/fuel-details/upload")
async def upload_fuel_details_workbook(file: UploadFile = File(...)) -> dict[str, Any]:
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 格式的 Excel 文件。")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空。")

    try:
        buffer = BytesIO(content)
        workbook = openpyxl.load_workbook(buffer, read_only=True, data_only=True)
        missing = {"加油明细"} - set(workbook.sheetnames)
        workbook.close()
        if missing:
            raise HTTPException(status_code=400, detail=f"Excel 缺少必要 sheet：{', '.join(sorted(missing))}")
        rows = read_fuel_details_from_workbook(BytesIO(content))
        return {
            "rows": rows,
            "source": "upload",
            "workbook": filename,
            "is_template": False,
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"无法读取上传的 Excel：{exc}") from exc


@app.post("/api/fuel-details/draft")
def save_fuel_details_draft(payload: FuelDetailsRequest) -> dict[str, Any]:
    data = {"rows": payload.rows}
    FUEL_DETAILS_DRAFT_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"saved": True, "count": len(payload.rows)}


@app.post("/api/fuel-details/export")
def export_fuel_details(payload: FuelDetailsRequest) -> StreamingResponse:
    output_name = payload.output_name or f"加油明细_维护导出_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    stream = BytesIO()
    write_fuel_details(stream, payload.rows, active_workbook_path())
    return excel_download_response(stream, output_name)


@app.get("/api/destination-pool")
def destination_pool() -> dict[str, Any]:
    return {"destinations": read_destination_pool()}


@app.post("/api/destination-pool")
def update_destination_pool(payload: DestinationPoolRequest) -> dict[str, Any]:
    save_destination_pool(payload.destinations)
    return {"saved": True, "count": len(payload.destinations)}


@app.post("/api/export")
def export(payload: ExportRequest) -> StreamingResponse:
    if not payload.records:
        raise HTTPException(status_code=400, detail="没有可导出的明细记录。")

    trips = trips_from_records(payload.records)

    output_name = payload.output_name or f"加油明细_在线编辑器生成_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    stream = BytesIO()
    write_usage_sheet(active_workbook_path(), stream, trips, FUEL_RATE)
    return excel_download_response(stream, output_name)


@app.post("/api/export/full")
def export_full(payload: CombinedExportRequest) -> StreamingResponse:
    if not payload.records and not payload.fuel_details:
        raise HTTPException(status_code=400, detail="没有可导出的车辆使用明细或加油明细。")

    workbook_path = active_workbook_path()
    workbook = openpyxl.load_workbook(workbook_path)
    template_workbook = openpyxl.load_workbook(workbook_path)
    populate_usage_sheet(workbook, template_workbook, trips_from_records(payload.records), FUEL_RATE)
    populate_fuel_details_sheet(workbook, payload.fuel_details)

    output_name = payload.output_name or f"行程燃油完整备份_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    stream = BytesIO()
    workbook.save(stream)
    return excel_download_response(stream, output_name)


app.mount("/web", StaticFiles(directory=WEB_DIR), name="web")


if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "8000"))
    uvicorn.run("app:app", host=host, port=port, reload=False)
