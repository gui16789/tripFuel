#!/usr/bin/env python3
"""Generate the vehicle usage detail sheet from manually curated trips.

MVP workflow:
1. Edit vehicle_usage_destinations.csv.
2. Set AMAP_KEY to enable automatic driving distance lookup.
3. Run this script to create a new workbook with the generated sheet.

When distance_km is filled in the CSV, the script can run without AMAP_KEY.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from urllib.request import urlopen

import openpyxl
from openpyxl.styles import Alignment, Font


DEFAULT_INPUT = "加油明细.xlsx"
DEFAULT_OUTPUT = "加油明细_mvp生成.xlsx"
DEFAULT_DESTINATIONS = "vehicle_usage_destinations.csv"
DEFAULT_FUEL_PRICES = "fuel_prices_changchun_92.csv"
DEFAULT_ORIGIN = "安盟财产保险有限公司长春中心支公司"
DEFAULT_CITY = "长春"
DEFAULT_FUEL_RATE = 0.09

GEOCODE_URL = "https://restapi.amap.com/v3/geocode/geo"
DRIVING_URL = "https://restapi.amap.com/v3/direction/driving"


@dataclass
class Trip:
    date: datetime
    origin: str
    destination: str
    fuel_price: float | None
    distance_km: float | None = None


class AmapClient:
    def __init__(
        self,
        key: str,
        city: str,
        cache_path: Path,
        sleep_seconds: float = 0.15,
    ) -> None:
        self.key = key
        self.city = city
        self.cache_path = cache_path
        self.sleep_seconds = sleep_seconds
        self.cache: dict[str, Any] = {}
        if cache_path.exists():
            self.cache = json.loads(cache_path.read_text(encoding="utf-8"))

    def save(self) -> None:
        self.cache_path.write_text(
            json.dumps(self.cache, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _get_json(self, url: str, params: dict[str, Any]) -> dict[str, Any]:
        full_url = f"{url}?{urlencode(params)}"
        with urlopen(full_url, timeout=20) as response:
            data = json.loads(response.read().decode("utf-8"))
        if str(data.get("status")) != "1":
            raise RuntimeError(f"AMap request failed: {data.get('info') or data}")
        time.sleep(self.sleep_seconds)
        return data

    def geocode(self, address: str) -> str:
        cache_key = f"geocode::{self.city}::{address}"
        if cache_key in self.cache:
            return self.cache[cache_key]

        data = self._get_json(
            GEOCODE_URL,
            {
                "key": self.key,
                "address": address,
                "city": self.city,
                "output": "JSON",
            },
        )
        geocodes = data.get("geocodes") or []
        if not geocodes:
            raise RuntimeError(f"AMap cannot geocode address: {address}")

        location = geocodes[0].get("location")
        if not location:
            raise RuntimeError(f"AMap returned empty location for: {address}")

        self.cache[cache_key] = location
        return location

    def driving_distance_m(self, origin: str, stops: list[str]) -> int:
        if not stops:
            raise ValueError("Trip must contain at least one destination stop.")

        cache_key = f"drive-roundtrip::{self.city}::{origin}::{'|'.join(stops)}"
        if cache_key in self.cache:
            return int(self.cache[cache_key])

        origin_location = self.geocode(origin)
        stop_locations = [self.geocode(stop) for stop in stops]
        destination_location = stop_locations[-1]
        waypoints = ";".join(stop_locations[:-1])

        params: dict[str, Any] = {
            "key": self.key,
            "origin": origin_location,
            "destination": destination_location,
            "strategy": 0,
            "extensions": "base",
            "output": "JSON",
        }
        if waypoints:
            params["waypoints"] = waypoints

        outbound_data = self._get_json(DRIVING_URL, params)
        outbound_paths = ((outbound_data.get("route") or {}).get("paths")) or []
        if not outbound_paths:
            raise RuntimeError(f"AMap returned no driving path: {origin} -> {stops}")

        return_data = self._get_json(
            DRIVING_URL,
            {
                "key": self.key,
                "origin": destination_location,
                "destination": origin_location,
                "strategy": 0,
                "extensions": "base",
                "output": "JSON",
            },
        )
        return_paths = ((return_data.get("route") or {}).get("paths")) or []
        if not return_paths:
            raise RuntimeError(f"AMap returned no return driving path: {stops[-1]} -> {origin}")

        distance_m = int(float(outbound_paths[0]["distance"])) + int(float(return_paths[0]["distance"]))
        self.cache[cache_key] = distance_m
        return distance_m


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate 车辆使用明细表 from manually curated destinations.",
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Source workbook path.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output workbook path.")
    parser.add_argument(
        "--destinations",
        default=DEFAULT_DESTINATIONS,
        help="CSV with date,destination,fuel_price,distance_km columns.",
    )
    parser.add_argument(
        "--source",
        choices=["csv", "sheet"],
        default="csv",
        help="Read trips from CSV or from the source workbook's 车辆使用明细表.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Only process the first N trips. Useful for AMap API smoke tests.",
    )
    parser.add_argument(
        "--origin",
        default=DEFAULT_ORIGIN,
        help="Default trip origin when CSV origin is empty.",
    )
    parser.add_argument("--city", default=DEFAULT_CITY, help="AMap city hint.")
    parser.add_argument(
        "--fuel-rate",
        type=float,
        default=DEFAULT_FUEL_RATE,
        help="Fuel consumption per km, e.g. 0.09 means 9L/100km.",
    )
    parser.add_argument(
        "--amap-key",
        default=os.environ.get("AMAP_KEY", ""),
        help="AMap Web service key. Defaults to AMAP_KEY environment variable.",
    )
    parser.add_argument(
        "--cache",
        default=".amap_cache.json",
        help="Local geocode/distance cache path.",
    )
    parser.add_argument(
        "--fuel-prices",
        default=DEFAULT_FUEL_PRICES,
        help="CSV with effective_date,price columns for Changchun 92 gasoline.",
    )
    parser.add_argument(
        "--round-km",
        choices=["nearest", "ceil", "floor"],
        default="nearest",
        help="Round AMap distance to integer km.",
    )
    return parser.parse_args()


def parse_date(raw: str) -> datetime:
    raw = raw.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            pass
    raise ValueError(f"Unsupported date format: {raw}")


def date_key(value: datetime) -> datetime:
    return datetime(value.year, value.month, value.day)


def parse_float(raw: str | None) -> float | None:
    if raw is None or not raw.strip():
        return None
    return float(raw.strip())


def split_stops(destination: str) -> list[str]:
    stops = [part.strip() for part in re.split(r"\s*[-—－]\s*", destination) if part.strip()]
    return stops or [destination.strip()]


def round_distance_km(distance_m: int, mode: str) -> int:
    km = distance_m / 1000
    if mode == "ceil":
        return int(-(-distance_m // 1000))
    if mode == "floor":
        return int(distance_m // 1000)
    return int(round(km))


def read_trips(csv_path: Path, default_origin: str) -> list[Trip]:
    trips: list[Trip] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"date", "destination", "fuel_price"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"CSV missing required columns: {', '.join(sorted(missing))}")

        for row_number, row in enumerate(reader, start=2):
            if not any((value or "").strip() for value in row.values()):
                continue

            destination = (row.get("destination") or "").strip()
            if not destination:
                raise ValueError(f"Row {row_number}: destination is required.")

            fuel_price = parse_float(row.get("fuel_price"))

            trips.append(
                Trip(
                    date=parse_date(row.get("date") or ""),
                    origin=(row.get("origin") or "").strip() or default_origin,
                    destination=destination,
                    fuel_price=fuel_price,
                    distance_km=parse_float(row.get("distance_km")),
                )
            )
    return trips


def read_trips_from_sheet(workbook_path: Path, default_origin: str) -> list[Trip]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = workbook["车辆使用明细表"]
    trips: list[Trip] = []

    for row in range(3, ws.max_row + 1):
        serial_no = ws.cell(row, 1).value
        if not isinstance(serial_no, int):
            continue

        date = ws.cell(row, 2).value
        destination = ws.cell(row, 4).value
        fuel_price = ws.cell(row, 7).value
        if not isinstance(date, datetime) or not destination:
            continue

        distance_km = ws.cell(row, 5).value
        trips.append(
            Trip(
                date=date,
                origin=ws.cell(row, 3).value or default_origin,
                destination=str(destination).strip(),
                fuel_price=float(fuel_price) if isinstance(fuel_price, (int, float)) else None,
                distance_km=float(distance_km) if isinstance(distance_km, (int, float)) else None,
            )
        )

    return trips


def read_fuel_prices(price_path: Path) -> list[tuple[datetime, float]]:
    if not price_path.exists():
        return []

    prices: list[tuple[datetime, float]] = []
    with price_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"effective_date", "price"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"Fuel price CSV missing required columns: {', '.join(sorted(missing))}")

        for row_number, row in enumerate(reader, start=2):
            if not any((value or "").strip() for value in row.values()):
                continue

            price = parse_float(row.get("price"))
            if price is None:
                raise ValueError(f"Fuel price row {row_number}: price is required.")
            prices.append((date_key(parse_date(row.get("effective_date") or "")), price))

    return sorted(prices, key=lambda item: item[0])


def resolve_fuel_prices(trips: list[Trip], price_path: Path) -> None:
    missing = [trip for trip in trips if trip.fuel_price is None]
    if not missing:
        return

    prices = read_fuel_prices(price_path)
    if not prices:
        raise RuntimeError(
            "Some trips have empty fuel_price, but no fuel price table is available. "
            f"Create {price_path} with effective_date,price columns or fill fuel_price."
        )

    for trip in missing:
        applicable = [price for effective_date, price in prices if effective_date <= date_key(trip.date)]
        if not applicable:
            raise RuntimeError(
                f"No fuel price found on or before {trip.date:%Y-%m-%d}. "
                f"Add an earlier effective_date to {price_path}."
            )
        trip.fuel_price = applicable[-1]


def resolve_distances(
    trips: list[Trip],
    amap_key: str,
    city: str,
    cache_path: Path,
    round_km: str,
) -> None:
    missing = [trip for trip in trips if trip.distance_km is None]
    if not missing:
        return

    if not amap_key:
        raise RuntimeError(
            "Some CSV rows have empty distance_km. Set AMAP_KEY or fill distance_km "
            "for offline MVP validation."
        )

    client = AmapClient(amap_key, city, cache_path)
    try:
        for trip in missing:
            distance_m = client.driving_distance_m(trip.origin, split_stops(trip.destination))
            trip.distance_km = round_distance_km(distance_m, round_km)
    finally:
        client.save()


def copy_row_style(ws: openpyxl.worksheet.worksheet.Worksheet, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)


def clear_usage_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= 3:
            ws.unmerge_cells(str(merged))

    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)


def write_usage_sheet(
    workbook_path: Path,
    output_path: Path,
    trips: list[Trip],
    fuel_rate: float,
) -> None:
    workbook = openpyxl.load_workbook(workbook_path)
    ws = workbook["车辆使用明细表"]

    template_workbook = openpyxl.load_workbook(workbook_path)
    template_ws = template_workbook["车辆使用明细表"]

    clear_usage_rows(ws)

    data_start_row = 3
    for index, trip in enumerate(sorted(trips, key=lambda item: item.date), start=1):
        row = data_start_row + index - 1
        copy_row_style(template_ws, 3, row)

        distance_km = int(round(float(trip.distance_km or 0)))
        ws.cell(row, 1, index)
        ws.cell(row, 2, trip.date)
        ws.cell(row, 2).number_format = "yyyy-mm-dd"
        ws.cell(row, 3, trip.origin)
        ws.cell(row, 4, trip.destination)
        ws.cell(row, 5, distance_km)
        ws.cell(row, 6, f"=E{row}*{fuel_rate}")
        ws.cell(row, 7, float(trip.fuel_price or 0))
        ws.cell(row, 8, f"=F{row}*G{row}")
        ws.cell(row, 9, f"=F{row}/E{row}*100")

    total_row = data_start_row + len(trips)
    copy_row_style(template_ws, 45, total_row)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(total_row, 1, "合计")
    ws.cell(total_row, 5, f"=SUM(E{data_start_row}:E{total_row - 1})")
    ws.cell(total_row, 6, f"=SUM(F{data_start_row}:F{total_row - 1})")
    ws.cell(total_row, 8, f"=SUM(H{data_start_row}:H{total_row - 1})")
    ws.cell(total_row, 9, f"=SUM(I{data_start_row}:I{total_row - 1})")
    for column in range(1, 10):
        ws.cell(total_row, column).font = copy(template_ws.cell(45, column).font) if template_ws.max_row >= 45 else Font(bold=True)
        ws.cell(total_row, column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def print_summary(trips: list[Trip], fuel_rate: float) -> None:
    total_km = sum(float(trip.distance_km or 0) for trip in trips)
    total_liters = total_km * fuel_rate
    total_amount = sum(float(trip.distance_km or 0) * fuel_rate * float(trip.fuel_price or 0) for trip in trips)
    print(f"Trips: {len(trips)}")
    print(f"Total km: {total_km:.0f}")
    print(f"Total fuel liters: {total_liters:.2f}")
    print(f"Total fuel amount: {total_amount:.2f}")


def main() -> int:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    destinations_path = Path(args.destinations)

    if not input_path.exists():
        print(f"Input workbook does not exist: {input_path}", file=sys.stderr)
        return 1
    if args.source == "csv" and not destinations_path.exists():
        print(f"Destination CSV does not exist: {destinations_path}", file=sys.stderr)
        return 1

    if args.source == "sheet":
        trips = read_trips_from_sheet(input_path, args.origin)
    else:
        trips = read_trips(destinations_path, args.origin)

    if args.limit:
        trips = trips[: args.limit]

    resolve_fuel_prices(trips, Path(args.fuel_prices))
    resolve_distances(
        trips,
        amap_key=args.amap_key,
        city=args.city,
        cache_path=Path(args.cache),
        round_km=args.round_km,
    )
    write_usage_sheet(input_path, output_path, trips, args.fuel_rate)
    print_summary(trips, args.fuel_rate)
    print(f"Generated: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
